import json
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from rich.console import Console
from rich.prompt import Prompt
from database import supabase

console = Console()

def get_orders_by_date(date_str):
    """Busca pedidos por data específica."""
    try:
        # Converter string de data para formato YYYY-MM-DD
        selected_date = datetime.strptime(date_str, '%d/%m/%Y').strftime('%Y-%m-%d')
        
        # Definir período de início e fim do dia
        start_of_day = f"{selected_date} 00:00:00"
        end_of_day = f"{selected_date} 23:59:59"
        
        # Consulta no Supabase
        query = supabase.table('tb_pedido') \
            .select('*, tb_cliente(nm_usuario), tb_endereco(rua, bairro, referencia)') \
            .gte('dt_registro', start_of_day) \
            .lte('dt_registro', end_of_day) \
            .execute()
            
        if not query.data:
            # Tentar abordagem alternativa se não encontrar registros pelo timestamp completo
            # Buscar por data sem considerar o horário (apenas LIKE na data)
            query = supabase.table('tb_pedido') \
                .select('*, tb_cliente(nm_usuario), tb_endereco(rua, bairro, referencia)') \
                .like('dt_registro', f"{selected_date}%") \
                .execute()
                
            if not query.data:
                return []
            
        return query.data
    except Exception as e:
        console.print(f"[red]Erro ao buscar pedidos da data {date_str}: {str(e)}[/red]")
        return []

def process_orders(orders):
    """Processa os pedidos para o formato necessário para a planilha."""
    processed_orders = []
    
    # Inicializar contadores totais
    total_marmitas = {'P': 0, 'M': 0, 'G': 0}
    total_valor_marmitas = 0
    total_valor_bebidas = 0
    total_valor_adicionais = 0
    total_geral = 0
    
    for order in orders:
        try:
            # Carregar as marmitas do JSON
            pedido_json = json.loads(order['marmitas'])
            
            # Contagem de marmitas por tamanho
            marmitas_count = {'P': 0, 'M': 0, 'G': 0}
            marmitas_valor = 0
            bebidas_valor = 0
            adicionais_valor = 0
            
            # Contar marmitas por tamanho
            if 'marmitas' in pedido_json:
                for marmita in pedido_json['marmitas']:
                    if isinstance(marmita, dict):
                        tamanho = marmita.get('tamanho', '')
                        if tamanho in marmitas_count:
                            marmitas_count[tamanho] += 1
                            total_marmitas[tamanho] += 1
                            marmitas_valor += float(marmita.get('preco', 0))
                            
                        # Calcular valor de adicionais
                        if 'adicionais' in marmita and isinstance(marmita['adicionais'], list):
                            for adicional in marmita['adicionais']:
                                adicionais_valor += float(adicional.get('preco', 0))
            
            # Calcular valor das bebidas
            if 'bebidas' in pedido_json:
                for bebida in pedido_json['bebidas']:
                    if isinstance(bebida, dict):
                        quantidade = bebida.get('quantidade', 1)
                        preco = bebida.get('preco', 0)
                        bebidas_valor += float(preco) * int(quantidade)
            
            # Criar linha do pedido
            cliente_nome = order['tb_cliente']['nm_usuario'] if order['tb_cliente'] else "Cliente não identificado"
            
            # Corrige o formato da data que pode conter timezone
            # Pega apenas a parte da data (antes do T ou do espaço)
            data_registro = order['dt_registro']
            if 'T' in data_registro:
                data_registro = data_registro.split('T')[0]
            else:
                data_registro = data_registro.split(' ')[0]
                
            data_pedido = datetime.strptime(data_registro, '%Y-%m-%d').strftime('%d/%m/%Y')
            
            order_row = {
                'Nome': cliente_nome,
                'Data': data_pedido,
                'P': marmitas_count['P'],
                'M': marmitas_count['M'],
                'G': marmitas_count['G'],
                'Refrigerante': len(pedido_json.get('bebidas', [])),
                'Extra Bife/Ovo': len([a for m in pedido_json.get('marmitas', []) 
                                     for a in m.get('adicionais', [])]),
                'Total': float(order['preco_total']),
                'Status Pagamento': order['status_pagamento'],
                'Forma Pagamento': order['forma_pagamento']
            }
            
            # Atualizar totais
            total_valor_marmitas += marmitas_valor
            total_valor_bebidas += bebidas_valor
            total_valor_adicionais += adicionais_valor
            total_geral += float(order['preco_total'])
            
            processed_orders.append(order_row)
        except Exception as e:
            console.print(f"[red]Erro ao processar pedido: {str(e)}[/red]")
    
    return processed_orders, {
        'total_unidades': sum(total_marmitas.values()),
        'total_refrigerantes': total_valor_bebidas,
        'total_extras': total_valor_adicionais,
        'total_geral': total_geral
    }

def create_excel(orders, totals, target_date_str):
    """Cria planilha Excel com os pedidos."""
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    
    # Cores
    AZUL_CLARO = 'B3E0FF'
    VERDE_CLARO = 'C6EFCE'
    LARANJA_CLARO = 'FFE4B5'
    ROXO_CLARO = 'E6E6FA'
    CINZA_CLARO = 'F0F0F0'
    
    # Estilos
    title_font = Font(name='Arial', size=16, bold=True, color='000080')  # Azul escuro
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Azul médio
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Estilo para células alternadas
    row_fills = [
        PatternFill(start_color=CINZA_CLARO, end_color=CINZA_CLARO, fill_type='solid'),
        PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    ]
    
    # Título
    ws['D1'] = 'Vendas Marmitas da Néia'
    ws.merge_cells('D1:J2')
    title_cell = ws['D1']
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type='solid')
    
    # Data do relatório
    ws['D3'] = f'Data: {target_date_str}'
    ws['D3'].font = Font(name='Arial', size=12, bold=True)
    
    # Cabeçalhos
    headers = ['Nome', 'Data', 'P', 'M', 'G', 'Refrigerante', 'Extra Bife/Ovo', 'Total', 'Status Pagamento', 'Forma Pagamento']
    for col, header in enumerate(headers, start=4):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar validação de dados para Status de Pagamento
    status_validation = DataValidation(
        type="list",
        formula1='"PAGO,PENDENTE"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Status Inválido",
        error="Por favor, selecione PAGO ou PENDENTE",
        showDropDown=True
    )
    ws.add_data_validation(status_validation)
    
    # Adicionar validação de dados para Forma de Pagamento
    payment_validation = DataValidation(
        type="list",
        formula1='"PIX,DINHEIRO,CARTÃO,NÃO INFORMADO"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Forma de Pagamento Inválida",
        error="Por favor, selecione uma forma de pagamento válida",
        showDropDown=True
    )
    ws.add_data_validation(payment_validation)
    
    # Adicionar formatação condicional para Status de Pagamento
    # Regra para PAGO (verde claro para linha inteira e texto em verde escuro)
    green_text = CellIsRule(
        operator='equal',
        formula=['="PAGO"'],
        stopIfTrue=True,
        font=Font(color='006100', bold=True),
        fill=PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    )
    # Regra para PENDENTE (texto em vermelho)
    red_text = CellIsRule(
        operator='equal',
        formula=['="PENDENTE"'],
        stopIfTrue=True,
        font=Font(color='FF0000', bold=True)
    )
    
    # Dados
    row = 5
    last_row = row + len(orders)
    
    # Aplicar formatação condicional para todas as colunas
    for col in range(ord('D'), ord('M')+1):  # De D até M (todas as colunas de dados, incluindo forma de pagamento)
        col_letter = chr(col)
        status_range = f'{col_letter}{row}:{col_letter}{last_row}'
        ws.conditional_formatting.add(status_range, green_text)
        if col == ord('L'):  # Apenas a coluna de status recebe a formatação vermelha
            ws.conditional_formatting.add(status_range, red_text)
    
    for idx, order in enumerate(orders):
        # Alternar cores das linhas
        row_fill = row_fills[idx % 2]
        
        for col, key in enumerate(headers):
            cell = ws.cell(row=row, column=col+4)
            value = order.get(key, '')
            cell.value = value
            cell.border = border
            cell.fill = row_fill
            
            if key == 'Total':
                cell.number_format = '#,##0.00'
                cell.value = float(value)
            elif key in ['P', 'M', 'G', 'Refrigerante', 'Extra Bife/Ovo']:
                cell.number_format = '0'
            elif key == 'Status Pagamento':
                # Adicionar validação à célula
                status_validation.add(cell)
            elif key == 'Forma Pagamento':
                # Adicionar validação à célula
                payment_validation.add(cell)
            
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
    
    # Resumo Financeiro
    summary_start_row = 4
    summary_style = Font(name='Arial', size=11, bold=True)
    summary_fill = PatternFill(start_color=VERDE_CLARO, end_color=VERDE_CLARO, fill_type='solid')
    
    # Título do resumo
    ws['N3'] = 'Resumo do Dia'
    ws.merge_cells('N3:O3')
    ws['N3'].font = Font(name='Arial', size=14, bold=True)
    ws['N3'].fill = PatternFill(start_color=LARANJA_CLARO, end_color=LARANJA_CLARO, fill_type='solid')
    ws['N3'].alignment = Alignment(horizontal='center')
    
    # Dados do resumo
    summary_data = [
        ('Total Unid. Marmitas:', totals['total_unidades']),
        ('Total de Refrigerantes:', f"R$ {totals['total_refrigerantes']:.2f}"),
        ('Total Extras:', f"R$ {totals['total_extras']:.2f}"),
        ('TOTAL:', f"R$ {totals['total_geral']:.2f}")
    ]
    
    for idx, (label, value) in enumerate(summary_data):
        row = summary_start_row + idx
        # Label
        ws[f'N{row}'] = label
        ws[f'N{row}'].font = summary_style
        ws[f'N{row}'].fill = summary_fill
        # Value
        ws[f'O{row}'] = value
        ws[f'O{row}'].font = summary_style
        ws[f'O{row}'].fill = summary_fill
        if idx == len(summary_data) - 1:  # Total em vermelho
            ws[f'O{row}'].font = Font(name='Arial', size=11, bold=True, color='FF0000')
    
    # Ajustar largura das colunas
    column_widths = {
        'D': 30,  # Nome
        'E': 12,  # Data
        'F': 8,   # P
        'G': 8,   # M
        'H': 8,   # G
        'I': 12,  # Refrigerante
        'J': 15,  # Extra Bife/Ovo
        'K': 12,  # Total
        'L': 18,  # Status Pagamento
        'M': 18,  # Forma Pagamento
        'N': 25,  # Resumo labels
        'O': 15   # Resumo values
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Preparar pasta para salvar o arquivo
    import os
    
    # Extrair o mês da data
    date_obj = datetime.strptime(target_date_str, '%d/%m/%Y')
    month_number = date_obj.month
    target_date_format = date_obj.strftime('%Y%m%d')
    
    # Criar estrutura de pastas
    base_dir = 'planilhas'
    month_dir = os.path.join(base_dir, str(month_number))
    
    # Criar pastas se não existirem
    if not os.path.exists(base_dir):
        os.makedirs(base_dir)
    if not os.path.exists(month_dir):
        os.makedirs(month_dir)
    
    # Definir caminho completo do arquivo
    excel_file = f'pedidos_{target_date_format}.xlsx'
    full_path = os.path.join(month_dir, excel_file)
    
    # Salvar arquivo
    wb.save(full_path)
    return full_path

def export_spreadsheet():
    """Interface para exportar a planilha."""
    console.clear()
    console.print("[bold yellow]=== Exportar Planilha de Pedidos ===[/bold yellow]\n")
    
    # Mostrar data atual para facilitar
    hoje = datetime.now().strftime('%d/%m/%Y')
    console.print(f"[cyan]Data atual: {hoje}[/cyan]\n")
    
    # Solicitar a data desejada
    data_desejada = Prompt.ask("Digite a data desejada (DD/MM/YYYY)", default=hoje)
    
    try:
        # Validar o formato da data
        datetime.strptime(data_desejada, '%d/%m/%Y')
        
        console.print(f"\n[cyan]Buscando pedidos do dia {data_desejada}...[/cyan]")
        
        # Buscar pedidos
        pedidos = get_orders_by_date(data_desejada)
        
        if not pedidos:
            console.print(f"\n[yellow]Nenhum pedido encontrado para a data {data_desejada}![/yellow]")
            input("\nPressione Enter para continuar...")
            return
        
        console.print(f"\n[green]Encontrados {len(pedidos)} pedidos![/green]")
        console.print("\n[cyan]Processando dados e gerando planilha...[/cyan]")
        
        # Processar pedidos
        orders_processed, totals = process_orders(pedidos)
        
        # Criar planilha
        excel_file = create_excel(orders_processed, totals, data_desejada)
        
        console.print(f"\n[bold green]Planilha gerada com sucesso![/bold green]")
        console.print(f"[green]Arquivo: {excel_file}[/green]")
    except ValueError:
        console.print(f"\n[red]Formato de data inválido! Use o formato DD/MM/YYYY.[/red]")
    except Exception as e:
        console.print(f"\n[red]Erro ao gerar planilha: {str(e)}[/red]")
    
    input("\nPressione Enter para continuar...")

if __name__ == '__main__':
    export_spreadsheet()
