import json
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

def load_data(filename='marmitasdaneia-79eb2-default-rtdb-1-export.json'):
    with open(filename, 'r', encoding='utf-8') as file:
        return json.load(file)

def process_orders(data, target_date):
    orders = []
    
    # Inicializar contadores totais
    total_marmitas = {'P': 0, 'M': 0, 'G': 0}
    total_valor_marmitas = 0
    total_valor_bebidas = 0
    total_valor_adicionais = 0
    total_geral = 0
    
    # Processar pedidos
    try:
        for pedido in data:
            if not pedido or not isinstance(pedido, dict):
                continue
                
            if pedido.get('data') == target_date:
                # Contagem de marmitas por tamanho
                marmitas_count = {'P': 0, 'M': 0, 'G': 0}
                for marmita in pedido['marmitas']:
                    marmitas_count[marmita['tamanho']] += 1
                    total_marmitas[marmita['tamanho']] += 1
                
                # Criar linha do pedido
                order = {
                    'Nome': pedido['nome_cliente'],
                    'Data': datetime.strptime(pedido['data'], '%Y-%m-%d').strftime('%d/%m/%Y'),
                    'P': marmitas_count['P'],
                    'M': marmitas_count['M'],
                    'G': marmitas_count['G'],
                    'Refrigerante': pedido['precos'].get('bebidas', 0) / 5 if pedido['precos'].get('bebidas', 0) > 0 else 0,
                    'Extra Bife/Ovo': pedido['precos'].get('adicionais', 0) / 7 if pedido['precos'].get('adicionais', 0) > 0 else 0,
                    'Total': float(pedido['precos'].get('total', pedido['precos']['marmitas'])),
                    'Status Pagamento': pedido['pagamento']['status'],
                    'Forma Pagamento': pedido['pagamento'].get('forma', 'NÃO INFORMADO')
                }
                
                # Atualizar totais
                total_valor_marmitas += float(pedido['precos']['marmitas'])
                total_valor_bebidas += float(pedido['precos'].get('bebidas', 0))
                total_valor_adicionais += float(pedido['precos'].get('adicionais', 0))
                total_geral += float(pedido['precos'].get('total', pedido['precos']['marmitas']))
                
                orders.append(order)
    except Exception as e:
        print(f"Erro ao processar pedidos: {e}")
    
    return orders, {
        'total_unidades': sum(total_marmitas.values()),
        'total_refrigerantes': total_valor_bebidas,
        'total_extras': total_valor_adicionais,
        'total_geral': total_geral
    }

def create_excel(target_date):
    # Carregar dados
    data = load_data()
    
    # Processar pedidos
    orders, totals = process_orders(data, target_date)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    
    # Cores
    AZUL_CLARO = 'B3E0FF'
    VERDE_CLARO = 'C6EFCE'
    LARANJA_CLARO = 'FFE4B5'
    ROXO_CLARO = 'E6E6FA'
    CINZA_CLARO = 'F0F0F0'
    
    # Estilos
    title_font = Font(name='Arial', size=16, bold=True, color='000080')  # Azul escuro
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Azul médio
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Estilo para células alternadas
    row_fills = [
        PatternFill(start_color=CINZA_CLARO, end_color=CINZA_CLARO, fill_type='solid'),
        PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    ]
    
    # Título
    ws['D1'] = 'Vendas Marmitas da Néia'
    ws.merge_cells('D1:J2')
    title_cell = ws['D1']
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type='solid')
    
    # Cabeçalhos
    headers = ['Nome', 'Data', 'P', 'M', 'G', 'Refrigerante', 'Extra Bife/Ovo', 'Total', 'Status Pagamento', 'Forma Pagamento']
    for col, header in enumerate(headers, start=4):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar validação de dados para Status de Pagamento
    status_validation = DataValidation(
        type="list",
        formula1='"PAGO,PENDENTE"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Status Inválido",
        error="Por favor, selecione PAGO ou PENDENTE",
        showDropDown=True
    )
    ws.add_data_validation(status_validation)
    
    # Adicionar validação de dados para Forma de Pagamento
    payment_validation = DataValidation(
        type="list",
        formula1='"PIX,DINHEIRO,CARTÃO,NÃO INFORMADO"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Forma de Pagamento Inválida",
        error="Por favor, selecione uma forma de pagamento válida",
        showDropDown=True
    )
    ws.add_data_validation(payment_validation)
    
    # Adicionar formatação condicional para Status de Pagamento
    # Regra para PAGO (verde claro para linha inteira e texto em verde escuro)
    green_text = CellIsRule(
        operator='equal',
        formula=['="PAGO"'],
        stopIfTrue=True,
        font=Font(color='006100', bold=True),
        fill=PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    )
    # Regra para PENDENTE (texto em vermelho)
    red_text = CellIsRule(
        operator='equal',
        formula=['="PENDENTE"'],
        stopIfTrue=True,
        font=Font(color='FF0000', bold=True)
    )
    
    # Dados
    row = 5
    last_row = row + len(orders)
    
    # Aplicar formatação condicional para todas as colunas
    for col in range(ord('D'), ord('M')+1):  # De D até M (todas as colunas de dados, incluindo forma de pagamento)
        col_letter = chr(col)
        status_range = f'{col_letter}{row}:{col_letter}{last_row}'
        ws.conditional_formatting.add(status_range, green_text)
        if col == ord('L'):  # Apenas a coluna de status recebe a formatação vermelha
            ws.conditional_formatting.add(status_range, red_text)
    
    for idx, order in enumerate(orders):
        # Alternar cores das linhas
        row_fill = row_fills[idx % 2]
        
        for col, key in enumerate(headers):
            cell = ws.cell(row=row, column=col+4)
            value = order.get(key, '')
            cell.value = value
            cell.border = border
            cell.fill = row_fill
            
            if key == 'Total':
                cell.number_format = '#,##0.00'
                cell.value = float(value)
            elif key in ['P', 'M', 'G', 'Refrigerante', 'Extra Bife/Ovo']:
                cell.number_format = '0'
            elif key == 'Status Pagamento':
                # Adicionar validação à célula
                status_validation.add(cell)
            elif key == 'Forma Pagamento':
                # Adicionar validação à célula
                payment_validation.add(cell)
            
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
    
    # Resumo Financeiro
    summary_start_row = 4
    summary_style = Font(name='Arial', size=11, bold=True)
    summary_fill = PatternFill(start_color=VERDE_CLARO, end_color=VERDE_CLARO, fill_type='solid')
    
    # Título do resumo
    ws['N3'] = 'Resumo do Dia'
    ws.merge_cells('N3:O3')
    ws['N3'].font = Font(name='Arial', size=14, bold=True)
    ws['N3'].fill = PatternFill(start_color=LARANJA_CLARO, end_color=LARANJA_CLARO, fill_type='solid')
    ws['N3'].alignment = Alignment(horizontal='center')
    
    # Dados do resumo
    summary_data = [
        ('Total Unid. Marmitas:', totals['total_unidades']),
        ('Total de Refrigerantes:', f"R$ {totals['total_refrigerantes']:.2f}"),
        ('Total Extras:', f"R$ {totals['total_extras']:.2f}"),
        ('TOTAL:', f"R$ {totals['total_geral']:.2f}")
    ]
    
    for idx, (label, value) in enumerate(summary_data):
        row = summary_start_row + idx
        # Label
        ws[f'N{row}'] = label
        ws[f'N{row}'].font = summary_style
        ws[f'N{row}'].fill = summary_fill
        # Value
        ws[f'O{row}'] = value
        ws[f'O{row}'].font = summary_style
        ws[f'O{row}'].fill = summary_fill
        if idx == len(summary_data) - 1:  # Total em vermelho
            ws[f'O{row}'].font = Font(name='Arial', size=11, bold=True, color='FF0000')
    
    # Ajustar largura das colunas
    column_widths = {
        'D': 30,  # Nome
        'E': 12,  # Data
        'F': 8,   # P
        'G': 8,   # M
        'H': 8,   # G
        'I': 12,  # Refrigerante
        'J': 15,  # Extra Bife/Ovo
        'K': 12,  # Total
        'L': 18,  # Status Pagamento
        'M': 18,  # Forma Pagamento
        'N': 25,  # Resumo labels
        'O': 15   # Resumo values
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Salvar arquivo
    excel_file = f'pedidos_{target_date}.xlsx'
    wb.save(excel_file)
    return excel_file

if __name__ == '__main__':
    target_date = '2025-03-01'
    excel_file = create_excel(target_date)
    print(f'Planilha gerada com sucesso: {excel_file}')
